from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
import sys
import unicodedata
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
LOCAL_DATA_DIR = BASE_DIR / "data"
REPO_INPUT_DIR = BASE_DIR / "Data" / "input"
REPO_OUTPUT_DIR = BASE_DIR / "Data" / "output"
OUTPUT_FILE = (REPO_OUTPUT_DIR if REPO_OUTPUT_DIR.exists() else BASE_DIR) / "tong_hop_bao_cao.xlsx"
SHEET_TONG_HOP = "Tong hop"
CHUA_CAP_NHAT = "Ch\u01b0a c\u1eadp nh\u1eadt"


def default_data_dir() -> Path:
    if LOCAL_DATA_DIR.exists() and any(LOCAL_DATA_DIR.glob("*.xlsx")):
        return LOCAL_DATA_DIR
    return REPO_INPUT_DIR


DATA_DIR = default_data_dir()


@dataclass(frozen=True)
class DongDuLieu:
    ten_phan_xuong: str
    stt_nguon: Any
    ngay: date
    gio: time | None
    truong_ca: str
    noi_dung: str
    file_nguon: str


@dataclass(frozen=True)
class DongTongHop:
    ngay: date
    ten_phan_xuong: str
    gio: time | None
    truong_ca: str
    noi_dung: str


@dataclass(frozen=True)
class DongBaoCao:
    stt: int
    ngay: date
    ten_phan_xuong: str
    noi_dung: str


def remove_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = remove_accents(text).strip().lower()
    return " ".join(text.split())


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value).strip()
    return text or CHUA_CAP_NHAT


def parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def parse_time(value: Any) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    text = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            pass
    return None


def time_sort_key(value: time | None) -> str:
    return value.strftime("%H:%M:%S") if value else "99:99:99"


def time_to_text(value: time | None) -> str:
    return value.strftime("%H:%M") if value else ""


def build_header_map(headers: tuple[Any, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, header in enumerate(headers):
        name = normalize_header(header)
        if name:
            header_map[name] = index
    return header_map


def find_col(header_map: dict[str, int], *names: str) -> int | None:
    for name in names:
        index = header_map.get(normalize_header(name))
        if index is not None:
            return index
    return None


def buoc_1_doc_du_lieu(data_dir: Path | None = None) -> list[DongDuLieu]:
    resolved_data_dir = (data_dir or DATA_DIR).resolve()
    if not resolved_data_dir.exists():
        raise FileNotFoundError(f"Khong tim thay thu muc du lieu: {resolved_data_dir}")

    rows: list[DongDuLieu] = []
    files = sorted(resolved_data_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"Khong tim thay file .xlsx nao trong thu muc: {data_dir}")

    for file_path in files:
        workbook = load_workbook(file_path, data_only=True)
        worksheet = workbook.active
        all_rows = list(worksheet.iter_rows(values_only=True))
        if not all_rows:
            continue

        header_map = build_header_map(all_rows[0])
        stt_col = find_col(header_map, "stt")
        ngay_col = find_col(header_map, "ngay")
        gio_col = find_col(header_map, "gio", "thoi gian")
        truong_ca_col = find_col(header_map, "truong ca")
        noi_dung_col = find_col(header_map, "noi dung")

        if ngay_col is None:
            raise ValueError(f"File {file_path.name} thieu cot Ngay")

        ten_phan_xuong = file_path.stem
        for excel_row in all_rows[1:]:
            ngay = parse_date(excel_row[ngay_col] if ngay_col < len(excel_row) else None)
            if ngay is None:
                continue

            rows.append(
                DongDuLieu(
                    ten_phan_xuong=ten_phan_xuong,
                    stt_nguon=excel_row[stt_col] if stt_col is not None and stt_col < len(excel_row) else "",
                    ngay=ngay,
                    gio=parse_time(excel_row[gio_col] if gio_col is not None and gio_col < len(excel_row) else None),
                    truong_ca=normalize_text(
                        excel_row[truong_ca_col] if truong_ca_col is not None and truong_ca_col < len(excel_row) else None
                    ),
                    noi_dung=normalize_text(
                        excel_row[noi_dung_col] if noi_dung_col is not None and noi_dung_col < len(excel_row) else None
                    ),
                    file_nguon=file_path.name,
                )
            )

    return rows


def buoc_2_tong_hop_theo_ngay(rows: list[DongDuLieu]) -> list[DongTongHop]:
    tong_hop = [
        DongTongHop(
            ngay=row.ngay,
            ten_phan_xuong=row.ten_phan_xuong,
            gio=row.gio,
            truong_ca=row.truong_ca,
            noi_dung=row.noi_dung,
        )
        for row in rows
    ]
    return sorted(tong_hop, key=lambda row: (row.ngay, row.ten_phan_xuong, time_sort_key(row.gio)))


def buoc_3_tao_du_lieu_bao_cao(rows: list[DongTongHop]) -> list[DongBaoCao]:
    grouped: dict[tuple[date, str], list[DongTongHop]] = defaultdict(list)
    for row in rows:
        grouped[(row.ngay, row.ten_phan_xuong)].append(row)

    bao_cao: list[DongBaoCao] = []
    label_truong_ca = "Tr\u01b0\u1edfng ca"
    label_noi_dung = "N\u1ed9i dung b\u00e1o c\u00e1o"

    for stt, ((ngay, ten_phan_xuong), items) in enumerate(sorted(grouped.items()), start=1):
        noi_dung_items = []
        for item in sorted(items, key=lambda row: time_sort_key(row.gio)):
            prefix = f"{time_to_text(item.gio)} - " if item.gio else ""
            noi_dung_items.append(f"{prefix}{label_truong_ca}: {item.truong_ca}. {label_noi_dung}: {item.noi_dung}")
        bao_cao.append(
            DongBaoCao(
                stt=stt,
                ngay=ngay,
                ten_phan_xuong=ten_phan_xuong,
                noi_dung="\n".join(noi_dung_items),
            )
        )
    return bao_cao


def write_rows(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    format_sheet(worksheet)


def format_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def tao_excel(
    du_lieu_buoc_1: list[DongDuLieu],
    du_lieu_buoc_2: list[DongTongHop],
    du_lieu_buoc_3: list[DongBaoCao],
    output_file: Path = OUTPUT_FILE,
) -> None:
    output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_TONG_HOP

    write_rows(
        worksheet,
        ["STT", "Ng\u00e0y", "T\u00ean ph\u00e2n x\u01b0\u1edfng", "N\u1ed9i dung"],
        [[row.stt, row.ngay, row.ten_phan_xuong, row.noi_dung] for row in du_lieu_buoc_3],
    )

    worksheet_step_1 = workbook.create_sheet("Buoc 1 - Du lieu doc")
    write_rows(
        worksheet_step_1,
        [
            "T\u00ean ph\u00e2n x\u01b0\u1edfng",
            "STT ngu\u1ed3n",
            "Ng\u00e0y",
            "Gi\u1edd",
            "Tr\u01b0\u1edfng ca",
            "N\u1ed9i dung",
            "File ngu\u1ed3n",
        ],
        [
            [
                row.ten_phan_xuong,
                row.stt_nguon,
                row.ngay,
                time_to_text(row.gio),
                row.truong_ca,
                row.noi_dung,
                row.file_nguon,
            ]
            for row in du_lieu_buoc_1
        ],
    )

    worksheet_step_2 = workbook.create_sheet("Buoc 2 - Tong hop ngay")
    write_rows(
        worksheet_step_2,
        ["Ng\u00e0y", "T\u00ean ph\u00e2n x\u01b0\u1edfng", "Gi\u1edd", "Tr\u01b0\u1edfng ca", "N\u1ed9i dung"],
        [[row.ngay, row.ten_phan_xuong, time_to_text(row.gio), row.truong_ca, row.noi_dung] for row in du_lieu_buoc_2],
    )

    workbook.save(output_file)


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    du_lieu_buoc_1 = buoc_1_doc_du_lieu()
    du_lieu_buoc_2 = buoc_2_tong_hop_theo_ngay(du_lieu_buoc_1)
    du_lieu_buoc_3 = buoc_3_tao_du_lieu_bao_cao(du_lieu_buoc_2)
    tao_excel(du_lieu_buoc_1, du_lieu_buoc_2, du_lieu_buoc_3)

    print(f"Read {len(du_lieu_buoc_1)} source rows.")
    print(f"Created {len(du_lieu_buoc_2)} step-2 summary rows.")
    print(f"Created {len(du_lieu_buoc_3)} final report rows in sheet '{SHEET_TONG_HOP}'.")
    print(f"Input dir: {DATA_DIR.resolve()}")
    print(f"Output file: {OUTPUT_FILE.resolve()}")


if __name__ == "__main__":
    main()
